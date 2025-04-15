import cv2
import numpy as np
import face_recognition
import os
import csv
import pickle
from datetime import datetime
import openpyxl

# Paths
path = 'faces'
encoding_file = 'encodings.pkl'
csv_filename = "attendance.csv"
excel_filename = "attendance.xlsx"

def save_encodings(encodings, names):
    with open(encoding_file, 'wb') as f:
        pickle.dump((encodings, names), f)

def load_encodings():
    if os.path.exists(encoding_file):
        with open(encoding_file, 'rb') as f:
            return pickle.load(f)
    return None, None

def encode_faces(images, names):
    encodings = []
    valid_names = []
    for img, name in zip(images, names):
        img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        faces = face_recognition.face_encodings(img_rgb)
        if faces:
            encodings.append(faces[0])
            valid_names.append(name)
    return encodings, valid_names

# Load encodings or process faces
known_encodings, classNames = load_encodings()

if known_encodings is None or classNames is None:
    print("🔍 Encoding faces for the first time...")
    images, names = [], []

    for file in os.listdir(path):
        img_path = os.path.join(path, file)
        img = cv2.imread(img_path)
        if img is not None:
            images.append(img)
            names.append(os.path.splitext(file)[0])

    known_encodings, classNames = encode_faces(images, names)
    save_encodings(known_encodings, classNames)
    print("✅ Encodings Saved!")
else:
    print("✅ Encodings Loaded Successfully!")

# Ensure CSV file has headers
if not os.path.exists(csv_filename):
    with open(csv_filename, 'w', newline='') as f:
        csv.writer(f).writerow(["Name", "Date", "Time"])

# Attendance tracking
marked_today = set()

def mark_attendance(name):
    now = datetime.now()
    date_today = now.strftime("%d-%m-%Y")
    time_now = now.strftime("%H:%M:%S")

    if os.path.exists(csv_filename):
        with open(csv_filename, 'r') as f:
            entries = f.readlines()
    else:
        entries = []

    if f"{name},{date_today}" not in "".join(entries):
        with open(csv_filename, 'a', newline='') as f:
            csv.writer(f).writerow([name, date_today, time_now])
        print(f"✅ {name}'s attendance marked!")
        marked_today.add(name)
    else:
        print(f"⚠ {name}'s attendance already marked today!")

def csv_to_excel():
    if not os.path.exists(csv_filename):
        return

    wb = openpyxl.Workbook() if not os.path.exists(excel_filename) else openpyxl.load_workbook(excel_filename)
    ws = wb.active
    ws.title = "Attendance"
    ws.delete_rows(1, ws.max_row)

    with open(csv_filename, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)

    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    wb.save(excel_filename)

# Live Video Feed
cap = cv2.VideoCapture(0)

while True:
    success, img = cap.read()
    if not success:
        break

    small_img = cv2.resize(img, (0, 0), fx=0.25, fy=0.25)
    rgb_small_img = cv2.cvtColor(small_img, cv2.COLOR_BGR2RGB)

    face_locations = face_recognition.face_locations(rgb_small_img)
    current_encodings = face_recognition.face_encodings(rgb_small_img, face_locations)

    for encodeFace, faceLoc in zip(current_encodings, face_locations):
        matches = face_recognition.compare_faces(known_encodings, encodeFace, tolerance=0.45)
        face_distances = face_recognition.face_distance(known_encodings, encodeFace)

        if len(face_distances) > 0:
            best_match_idx = np.argmin(face_distances)
            if matches[best_match_idx]:
                name = classNames[best_match_idx].upper()

                if name not in marked_today:
                    mark_attendance(name)

                # Draw bounding box
                y1, x2, y2, x1 = [v * 4 for v in faceLoc]
                cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                cv2.putText(img, name, (x1 + 6, y2 - 6),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255, 255, 255), 2)

    cv2.imshow("Face Attendance", img)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()

csv_to_excel()
os.remove(csv_filename)
